from dotenv import load_dotenv
from pathlib import Path
import os

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import logging
import uuid
import json
import base64
import secrets
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Annotated, Any

import bcrypt
import jwt
import requests
import httpx
from bson import ObjectId
from fastapi import FastAPI, APIRouter, Request, HTTPException, Depends, UploadFile, File, Header, Query
from fastapi.responses import Response, StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, BeforeValidator, EmailStr

from checklist_data import get_checklist, SECTIONS
from pdf_report import build_sir_pdf, build_ecr_pdf, build_hor_pdf, std_filename
from email_service import send_email, spare_change_html, daily_summary_html, approval_html

# tambahan untuk upload foto mulai disini
from fastapi.responses import FileResponse
# tambahan untuk upload foto sampai disini

# ---------------------------------------------------------------------------
# Setup
# ---------------------------------------------------------------------------
mongo_url = os.getenv('MONGO_URL', 'mongodb+srv://agussupriyadi_db_user:hMWD6WLIndqiecK2@cluster0.oe06guo.mongodb.net/?appName=Cluster0')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
scheduler = AsyncIOScheduler()

app = FastAPI(title="Fujitec SIR")
api_router = APIRouter(prefix="/api")

JWT_ALGORITHM = "HS256"
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_NAME = "fujitec-sir"
storage_key = None

ROLES = ["admin", "technician", "troubleshooter", "supervisor", "head_maintenance",
         "admin_staff", "contract_staff", "sales", "inventory", "customer"]
OWNER_EMAIL = os.environ.get("OWNER_EMAIL", "").lower()
AUTH_SESSION_URL = "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data"

MIME_TYPES = {
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
    "gif": "image/gif", "webp": "image/webp", "heic": "image/heic",
}

# ---------------------------------------------------------------------------
# Mongo helpers
# ---------------------------------------------------------------------------
PyObjectId = Annotated[str, BeforeValidator(str)]


def now_iso():
    return datetime.now(timezone.utc).isoformat()


# ---------------------------------------------------------------------------
# Storage
# ---------------------------------------------------------------------------
def init_storage():
    global storage_key
    if storage_key:
        return storage_key
    resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
    resp.raise_for_status()
    storage_key = resp.json()["storage_key"]
    return storage_key

# diganti dari sini ----
# def put_object(path: str, data: bytes, content_type: str) -> dict:
#    key = init_storage()
#    resp = requests.put(f"{STORAGE_URL}/objects/{path}",
#                        headers={"X-Storage-Key": key, "Content-Type": content_type},
#                        data=data, timeout=120)
#    resp.raise_for_status()
#    return resp.json()
#
# def get_object(path: str):
#    key = init_storage()
#    resp = requests.get(f"{STORAGE_URL}/objects/{path}",
#                        headers={"X-Storage-Key": key}, timeout=60)
#    resp.raise_for_status()
#    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")
#
# diganti sampai sini -----

# pengganti dari sini -----
# Folder lokal untuk menampung foto upload di server
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

def put_object(path: str, data: bytes, content_type: str = "image/jpeg"):
    """Menyimpan file foto ke disk lokal server alih-alih EmergentAgent"""
    try:
        # Buat nama file aman dari path
        safe_filename = path.replace("/", "_")
        filepath = os.path.join(UPLOAD_DIR, safe_filename)
        
        with open(filepath, "wb") as f:
            f.write(data)
            
        # Kembalikan URL publik / identifier file
        return {
            "path": path,
            "url": f"/api/files/{safe_filename}",
            "filename": safe_filename
        }
    except Exception as e:
        print(f"Error saving file locally: {e}")
        raise e

def get_object(path: str):
    """Membaca file foto dari disk lokal server"""
    safe_filename = path.replace("/", "_")
    filepath = os.path.join(UPLOAD_DIR, safe_filename)
    if os.path.exists(filepath):
        with open(filepath, "rb") as f:
            return f.read()
    return None
# pengganti sampai sini -----

# ---------------------------------------------------------------------------
# Auth utilities
# ---------------------------------------------------------------------------
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))


def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]


def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {"sub": user_id, "email": email, "role": role,
               "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "access"}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["id"] = str(user["_id"])
        user.pop("_id", None)
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


def require_roles(*roles):
    async def checker(user: dict = Depends(get_current_user)):
        if user["role"] not in roles and user["role"] != "admin":
            raise HTTPException(status_code=403, detail="Not authorized for this action")
        return user
    return checker


def create_pdf_token(resource_id: str) -> str:
    payload = {"sub": resource_id, "type": "pdf",
               "exp": datetime.now(timezone.utc) + timedelta(minutes=15)}
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


def require_pdf_token(auth: str, resource_id: str):
    if not auth:
        raise HTTPException(status_code=401, detail="PDF token required")
    try:
        p = jwt.decode(auth, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid or expired PDF token")
    if p.get("type") != "pdf" or p.get("sub") != resource_id:
        raise HTTPException(status_code=401, detail="Invalid PDF token")


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class RegisterInput(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = "technician"


class LoginInput(BaseModel):
    email: EmailStr
    password: str


class Point(BaseModel):
    index: int
    measurement: str = ""
    photo_file_id: Optional[str] = None


class ChecklistItemResult(BaseModel):
    no: int
    section: str
    description: str
    photo_points: int = 0
    judgment: Optional[str] = None  # good | replaced | damage | none
    remark: str = ""
    points: List[Point] = []


class SparePart(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    quantity: str = ""
    sales_status: str = "pending"        # pending | offered
    customer_po_status: str = "pending"  # pending | po_issued
    inventory_status: str = "pending"    # pending | available | ordering | no_stock | discontinued
    maintenance_status: str = "pending"  # pending | replaced
    after_photo_file_id: Optional[str] = None  # proof photo after replacement
    note: str = ""


class Signature(BaseModel):
    name: str = ""
    image: str = ""  # data URL
    signed_at: Optional[str] = None


class Signatures(BaseModel):
    issued_by: Signature = Signature()
    customer: Signature = Signature()
    checked_by: Signature = Signature()
    approved_by: Signature = Signature()


class InspectionInput(BaseModel):
    job_number: str
    site_name: str
    inspection_date: str
    time_from: str
    time_to: str
    serviced_by: str
    total_units: int = 0
    lift_number: str
    checklist_type: str  # 1M | 3M | 12M
    items: List[ChecklistItemResult] = []
    global_remark: str = ""
    spare_parts: List[SparePart] = []
    signatures: Signatures = Signatures()
    customer_email: str = ""
    status: str = "draft"  # draft | submitted | checked | approved


class SparePartStatusUpdate(BaseModel):
    field: str  # sales_status | customer_po_status | inventory_status | maintenance_status | note
    value: str


# ---------------------------------------------------------------------------
# Auth endpoints
# ---------------------------------------------------------------------------
@api_router.post("/auth/register")
async def register(data: RegisterInput):
    email = data.email.lower()
    if data.role not in ROLES:
        raise HTTPException(status_code=400, detail="Invalid role")
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already registered")
    doc = {
        "email": email,
        "password_hash": hash_password(data.password),
        "name": data.name,
        "role": data.role,
        "created_at": now_iso(),
    }
    res = await db.users.insert_one(doc)
    uid = str(res.inserted_id)
    token = create_access_token(uid, email, data.role)
    return {"access_token": token, "user": {"id": uid, "email": email, "name": data.name, "role": data.role}}


@api_router.post("/auth/login")
async def login(data: LoginInput, request: Request):
    email = data.email.lower()
    ident = f"{request.client.host}:{email}"
    attempt = await db.login_attempts.find_one({"identifier": ident})
    if attempt and attempt.get("count", 0) >= 5:
        locked_until = attempt.get("locked_until")
        if locked_until and datetime.fromisoformat(locked_until) > datetime.now(timezone.utc):
            raise HTTPException(status_code=429, detail="Too many attempts. Try again later.")
    user = await db.users.find_one({"email": email})
    if not user or not user.get("password_hash") or not verify_password(data.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": ident},
            {"$inc": {"count": 1},
             "$set": {"locked_until": (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()}},
            upsert=True)
        raise HTTPException(status_code=401, detail="Invalid email or password")
    await db.login_attempts.delete_one({"identifier": ident})
    uid = str(user["_id"])
    token = create_access_token(uid, email, user["role"])
    return {"access_token": token, "user": {"id": uid, "email": email, "name": user["name"], "role": user["role"]}}


@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


class GoogleSessionInput(BaseModel):
    session_id: str


@api_router.post("/auth/google-session")
async def google_session(data: GoogleSessionInput):
    # Exchange Emergent session_id for profile, then mint our own JWT.
    try:
        async with httpx.AsyncClient(timeout=30) as hc:
            resp = await hc.get(AUTH_SESSION_URL, headers={"X-Session-ID": data.session_id})
        resp.raise_for_status()
        profile = resp.json()
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid or expired Google session")
    email = profile["email"].lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        uid = str(existing["_id"])
        role = existing["role"]
    else:
        role = "admin" if email == OWNER_EMAIL else "technician"
        doc = {"email": email, "name": profile.get("name", email), "role": role,
               "picture": profile.get("picture"), "auth_provider": "google", "created_at": now_iso()}
        res = await db.users.insert_one(doc)
        uid = str(res.inserted_id)
    token = create_access_token(uid, email, role)
    return {"access_token": token, "user": {"id": uid, "email": email,
            "name": profile.get("name", email), "role": role}}


@api_router.post("/auth/logout")
async def logout(user: dict = Depends(get_current_user)):
    return {"ok": True}


# ---------------------------------------------------------------------------
# Users (admin)
# ---------------------------------------------------------------------------
@api_router.get("/users")
async def list_users(user: dict = Depends(require_roles("admin", "supervisor", "head_maintenance"))):
    users = await db.users.find({}, {"password_hash": 0}).to_list(500)
    for u in users:
        u["id"] = str(u["_id"])
        u.pop("_id", None)
    return users


@api_router.post("/users")
async def create_user(data: RegisterInput, user: dict = Depends(require_roles("admin"))):
    return await register(data)


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, user: dict = Depends(require_roles("admin"))):
    await db.users.delete_one({"_id": ObjectId(user_id)})
    return {"ok": True}


# ---------------------------------------------------------------------------
# Checklist master
# ---------------------------------------------------------------------------
@api_router.get("/checklist/master")
async def checklist_master(type: str = Query(...), user: dict = Depends(get_current_user)):
    if type not in ("1M", "3M", "12M"):
        raise HTTPException(status_code=400, detail="Invalid checklist type")
    return {"type": type, "sections": SECTIONS, "items": get_checklist(type)}


# ---------------------------------------------------------------------------
# Buildings (master gedung)
# ---------------------------------------------------------------------------
@api_router.get("/buildings")
async def search_buildings(q: str = Query("", ), user: dict = Depends(get_current_user)):
    query = {}
    if q:
        query = {"$or": [
            {"job_number": {"$regex": q, "$options": "i"}},
            {"site_name": {"$regex": q, "$options": "i"}},
            {"city": {"$regex": q, "$options": "i"}},
        ]}
    docs = await db.buildings.find(query, {"_id": 0}).limit(20).to_list(20)
    return docs


@api_router.get("/buildings/lookup")
async def lookup_building(job_number: str = Query(...), user: dict = Depends(get_current_user)):
    doc = await db.buildings.find_one({"job_number": {"$regex": f"^{job_number}$", "$options": "i"}}, {"_id": 0})
    return doc or {}


# ---------------------------------------------------------------------------
# Files / photo upload
# ---------------------------------------------------------------------------
@api_router.post("/upload")
async def upload(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    ext = (file.filename.split(".")[-1] if "." in file.filename else "bin").lower()
    file_id = str(uuid.uuid4())
    path = f"{file_id}.{ext}"
    data = await file.read()
    ctype = file.content_type or MIME_TYPES.get(ext, "application/octet-stream")
    
    result = put_object(path, data, ctype)
    
    await db.files.insert_one({
        "id": file_id,
        "storage_path": result["filename"],
        "original_filename": file.filename,
        "content_type": ctype,
        "size": len(data),
        "uploaded_by": user["id"],
        "is_deleted": False,
        "created_at": now_iso(),
    })
    return {"file_id": file_id, "path": result["url"]}

@api_router.get("/files/{file_id}")
async def download_file(file_id: str):
    # Cari berdasarkan ID file di DB
    record = await db.files.find_one({"id": file_id, "is_deleted": False})
    
    if record:
        safe_filename = record.get("storage_path", "").split("/")[-1]
        filepath = os.path.join(UPLOAD_DIR, safe_filename)
        if os.path.exists(filepath):
            return FileResponse(filepath, media_type=record.get("content_type", "image/jpeg"))
            
    # Fallback jika dicari berdasarkan nama file langsung di folder uploads
    filepath = os.path.join(UPLOAD_DIR, file_id)
    if os.path.exists(filepath):
        return FileResponse(filepath, media_type="image/jpeg")

    raise HTTPException(status_code=404, detail="File not found")
        
    return Response(content=data, media_type=record.get("content_type", "image/jpeg"))

# changes dummy changes ini= this line

# ---------------------------------------------------------------------------
# Inspections
# ---------------------------------------------------------------------------
def _validate_complete(data: InspectionInput):
    incomplete = [it.no for it in data.items if not it.judgment]
    if incomplete:
        raise HTTPException(status_code=400,
            detail=f"Checklist belum 100% terisi. {len(incomplete)} item belum diberi judgment. Simpan sebagai draft dulu.")


def _visible_query(user: dict):
    """Restrict list results by role."""
    if user["role"] in ("admin", "supervisor", "head_maintenance", "sales", "inventory"):
        return {}
    if user["role"] == "technician":
        return {"technician_id": user["id"]}
    if user["role"] == "customer":
        return {}  # customers can see reports (could be scoped by site later)
    return {}


@api_router.post("/inspections")
async def create_inspection(data: InspectionInput, user: dict = Depends(get_current_user)):
    if data.status == "submitted":
        _validate_complete(data)
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["technician_id"] = user["id"]
    doc["technician_name"] = user["name"]
    doc["created_at"] = now_iso()
    doc["updated_at"] = now_iso()
    await db.inspections.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.get("/inspections")
async def list_inspections(user: dict = Depends(get_current_user),
                           status: Optional[str] = None,
                           checklist_type: Optional[str] = None,
                           site_name: Optional[str] = None,
                           q: Optional[str] = None):
    query = _visible_query(user)
    if status:
        query["status"] = status
    if checklist_type:
        query["checklist_type"] = checklist_type
    if site_name:
        query["site_name"] = site_name
    if q:
        query["$or"] = [
            {"job_number": {"$regex": q, "$options": "i"}},
            {"site_name": {"$regex": q, "$options": "i"}},
            {"lift_number": {"$regex": q, "$options": "i"}},
        ]
    docs = await db.inspections.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    # Trim heavy fields for list
    for d in docs:
        d.pop("items", None)
        d.pop("signatures", None)
    return docs


@api_router.get("/inspections/sites")
async def inspection_sites(user: dict = Depends(get_current_user)):
    query = _visible_query(user)
    sites = await db.inspections.distinct("site_name", query)
    return sorted([s for s in sites if s])


@api_router.get("/inspections/{inspection_id}")
async def get_inspection(inspection_id: str, user: dict = Depends(get_current_user)):
    doc = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Inspection not found")
    return doc


@api_router.put("/inspections/{inspection_id}")
async def update_inspection(inspection_id: str, data: InspectionInput, user: dict = Depends(get_current_user)):
    existing = await db.inspections.find_one({"id": inspection_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Inspection not found")
    # Editing rules: only draft reports may be edited, by owner or admin.
    is_owner = existing.get("technician_id") == user["id"]
    if existing.get("status") != "draft" and user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Laporan yang sudah dikirim tidak dapat diedit")
    if not is_owner and user["role"] not in ("admin",):
        raise HTTPException(status_code=403, detail="Hanya teknisi pembuat yang dapat mengedit draft ini")
    if data.status == "submitted":
        _validate_complete(data)
    doc = data.model_dump()
    doc["updated_at"] = now_iso()
    doc.pop("technician_id", None)
    doc.pop("technician_name", None)
    await db.inspections.update_one({"id": inspection_id}, {"$set": doc})
    updated = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    return updated


@api_router.patch("/inspections/{inspection_id}/status")
async def update_status(inspection_id: str, payload: dict, user: dict = Depends(get_current_user)):
    status = payload.get("status")
    if status not in ("draft", "submitted", "checked", "approved"):
        raise HTTPException(status_code=400, detail="Invalid status")
    existing = await db.inspections.find_one({"id": inspection_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Not found")
    current = existing.get("status", "draft")
    role = user["role"]

    def allowed():
        if status == "submitted":
            return role in ("technician", "supervisor", "head_maintenance", "admin")
        if status == "approved":
            return role in ("head_maintenance", "admin")
        if status in ("checked", "draft"):
            # downgrade from approved is restricted to head of maintenance
            if current == "approved":
                return role in ("head_maintenance", "admin")
            return role in ("supervisor", "head_maintenance", "admin")
        return False

    if not allowed():
        if current == "approved":
            raise HTTPException(status_code=403, detail="Hanya Kepala Maintenance yang dapat menurunkan status dari Approved")
        raise HTTPException(status_code=403, detail="Anda tidak berwenang mengubah status ini")
    await db.inspections.update_one({"id": inspection_id}, {"$set": {"status": status, "updated_at": now_iso()}})
    if status == "approved":
        try:
            fresh = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
            await _send_customer_pdf("SIR", fresh)
        except Exception as e:
            logger.error(f"customer email failed: {e}")
    return {"ok": True, "status": status}


@api_router.patch("/inspections/{inspection_id}/signature")
async def add_signature(inspection_id: str, payload: dict, user: dict = Depends(get_current_user)):
    key = payload.get("key")
    signature = payload.get("signature") or {}
    allowed = {
        "issued_by": ("technician", "supervisor", "head_maintenance", "admin"),
        "customer": tuple(ROLES),
        "checked_by": ("supervisor", "head_maintenance", "admin"),
        "approved_by": ("head_maintenance", "admin"),
    }
    if key not in allowed:
        raise HTTPException(status_code=400, detail="Invalid signature key")
    if user["role"] not in allowed[key]:
        raise HTTPException(status_code=403, detail="Anda tidak berwenang menandatangani bagian ini")
    signature.setdefault("signed_at", now_iso())
    if not signature.get("name"):
        signature["name"] = user["name"]
    res = await db.inspections.update_one(
        {"id": inspection_id}, {"$set": {f"signatures.{key}": signature, "updated_at": now_iso()}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


@api_router.delete("/inspections/{inspection_id}")
async def delete_inspection(inspection_id: str, user: dict = Depends(get_current_user)):
    doc = await db.inspections.find_one({"id": inspection_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    if user["role"] not in ("admin", "supervisor", "head_maintenance") and doc.get("technician_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    await db.inspections.delete_one({"id": inspection_id})
    return {"ok": True}


@api_router.get("/inspections/{inspection_id}/pdf")
async def inspection_pdf(inspection_id: str, auth: str = Query(None)):
    require_pdf_token(auth, inspection_id)
    doc = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    # Pre-fetch photos for the attachment sheet.
    photos = {}
    for it in doc.get("items", []):
        for p in it.get("points", []):
            fid = p.get("photo_file_id")
            if fid and fid not in photos:
                rec = await db.files.find_one({"id": fid, "is_deleted": False})
                if rec:
                    try:
                        data, _ = get_object(rec["storage_path"])
                        photos[fid] = data
                    except Exception:
                        pass
    pdf_bytes = build_sir_pdf(doc, photos=photos)
    filename = std_filename("SIR", doc.get("inspection_date"), doc.get("job_number"),
                            doc.get("site_name"), doc.get("lift_number"))
    return StreamingResponse(iter([pdf_bytes]), media_type="application/pdf",
                             headers={"Content-Disposition": f'inline; filename="{filename}"'})


# ---------------------------------------------------------------------------
# Spare parts follow-up
# ---------------------------------------------------------------------------
@api_router.get("/spare-parts")
async def list_spare_parts(user: dict = Depends(get_current_user),
                           inventory_status: Optional[str] = None,
                           only_open: bool = False):
    docs = await db.inspections.find(
        {"spare_parts.0": {"$exists": True}}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    rows = []
    for d in docs:
        for sp in d.get("spare_parts", []):
            row = dict(sp)
            row["inspection_id"] = d["id"]
            row["job_number"] = d.get("job_number")
            row["site_name"] = d.get("site_name")
            row["lift_number"] = d.get("lift_number")
            row["inspection_date"] = d.get("inspection_date")
            row["checklist_type"] = d.get("checklist_type")
            if inventory_status and row.get("inventory_status") != inventory_status:
                continue
            if only_open and row.get("maintenance_status") == "replaced":
                continue
            rows.append(row)
    return rows


@api_router.patch("/inspections/{inspection_id}/spare-parts/{spare_id}")
async def update_spare_part(inspection_id: str, spare_id: str,
                            update: SparePartStatusUpdate,
                            user: dict = Depends(get_current_user)):
    field = update.field
    allowed_by_role = {
        "sales_status": ("sales", "supervisor", "head_maintenance", "admin"),
        "customer_po_status": ("customer", "sales", "supervisor", "head_maintenance", "admin"),
        "inventory_status": ("inventory", "supervisor", "head_maintenance", "admin"),
        "maintenance_status": ("technician", "troubleshooter", "supervisor", "head_maintenance", "admin"),
        "after_photo_file_id": ("technician", "troubleshooter", "supervisor", "head_maintenance", "admin"),
        "note": tuple(ROLES),
    }
    if field not in allowed_by_role:
        raise HTTPException(status_code=400, detail="Invalid field")
    if user["role"] not in allowed_by_role[field]:
        raise HTTPException(status_code=403, detail="Not authorized for this field")
    res = await db.inspections.update_one(
        {"id": inspection_id, "spare_parts.id": spare_id},
        {"$set": {f"spare_parts.$.{field}": update.value, "updated_at": now_iso()}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Spare part not found")
    if field in ("sales_status", "customer_po_status", "inventory_status", "maintenance_status"):
        await _notify_spare_change(inspection_id, spare_id, field, update.value, user["name"])
    return {"ok": True}


SPARE_FIELD_LABELS = {
    "sales_status": "Sales / Penawaran", "customer_po_status": "Customer (PO)",
    "inventory_status": "Inventory", "maintenance_status": "Maintenance",
}
SPARE_VALUE_LABELS = {
    "pending": "Pending", "offered": "Sudah Ditawarkan", "po_issued": "PO Terbit",
    "available": "Tersedia", "ordering": "Sedang Dipesan", "no_stock": "Tidak Ada Stok",
    "discontinued": "Discontinue", "replaced": "Sudah Diganti",
}


async def _notify_spare_change(inspection_id, spare_id, field, value, changed_by):
    insp = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not insp:
        return
    sp = next((s for s in insp.get("spare_parts", []) if s.get("id") == spare_id), {})
    recipients = set()
    if OWNER_EMAIL:
        recipients.add(OWNER_EMAIL)
    async for u in db.users.find({"role": {"$in": ["head_maintenance", "supervisor"]}}, {"email": 1}):
        recipients.add(u["email"])
    html = spare_change_html(insp, sp, SPARE_FIELD_LABELS.get(field, field),
                             SPARE_VALUE_LABELS.get(value, value), changed_by)
    subject = f"[SIR] Update Spare Part — {sp.get('name','')} @ {insp.get('site_name','')}"
    for r in recipients:
        await send_email(r, subject, html)


# ---------------------------------------------------------------------------
# Reports: ECR (Call Back) & HOR (Hand Over)
# ---------------------------------------------------------------------------
REPORT_TYPES = ("ECR", "HOR")


@api_router.post("/reports")
async def create_report(payload: dict, user: dict = Depends(get_current_user)):
    rtype = payload.get("report_type")
    if rtype not in REPORT_TYPES:
        raise HTTPException(status_code=400, detail="Invalid report_type")
    payload["id"] = str(uuid.uuid4())
    payload["technician_id"] = user["id"]
    payload["technician_name"] = user["name"]
    payload.setdefault("status", "draft")
    payload["created_at"] = now_iso()
    payload["updated_at"] = now_iso()
    await db.reports.insert_one(payload)
    payload.pop("_id", None)
    return payload


@api_router.get("/reports")
async def list_reports(user: dict = Depends(get_current_user),
                       type: Optional[str] = None, status: Optional[str] = None, q: Optional[str] = None):
    query = {}
    if user["role"] == "technician":
        query["technician_id"] = user["id"]
    if type:
        query["report_type"] = type
    if status:
        query["status"] = status
    if q:
        query["$or"] = [
            {"site_name": {"$regex": q, "$options": "i"}},
            {"job_number": {"$regex": q, "$options": "i"}},
            {"unit_no": {"$regex": q, "$options": "i"}},
            {"lift_number": {"$regex": q, "$options": "i"}},
        ]
    docs = await db.reports.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return docs


@api_router.get("/reports/{report_id}")
async def get_report(report_id: str, user: dict = Depends(get_current_user)):
    doc = await db.reports.find_one({"id": report_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Report not found")
    return doc


@api_router.put("/reports/{report_id}")
async def update_report(report_id: str, payload: dict, user: dict = Depends(get_current_user)):
    existing = await db.reports.find_one({"id": report_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Report not found")
    is_owner = existing.get("technician_id") == user["id"]
    if existing.get("status") != "draft" and user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Laporan yang sudah dikirim tidak dapat diedit")
    if not is_owner and user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Hanya teknisi pembuat yang dapat mengedit draft ini")
    for k in ("id", "technician_id", "technician_name", "created_at", "report_type"):
        payload.pop(k, None)
    payload["updated_at"] = now_iso()
    await db.reports.update_one({"id": report_id}, {"$set": payload})
    return await db.reports.find_one({"id": report_id}, {"_id": 0})


@api_router.patch("/reports/{report_id}/status")
async def update_report_status(report_id: str, payload: dict, user: dict = Depends(get_current_user)):
    status = payload.get("status")
    if status not in ("draft", "submitted", "checked", "approved"):
        raise HTTPException(status_code=400, detail="Invalid status")
    existing = await db.reports.find_one({"id": report_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Not found")
    current = existing.get("status", "draft")
    role = user["role"]
    ok = (
        (status == "submitted" and role in ("technician", "supervisor", "head_maintenance", "admin")) or
        (status == "approved" and role in ("head_maintenance", "admin")) or
        (status in ("checked", "draft") and (
            (current == "approved" and role in ("head_maintenance", "admin")) or
            (current != "approved" and role in ("supervisor", "head_maintenance", "admin"))))
    )
    if not ok:
        raise HTTPException(status_code=403, detail="Anda tidak berwenang mengubah status ini")
    await db.reports.update_one({"id": report_id}, {"$set": {"status": status, "updated_at": now_iso()}})
    if status == "approved":
        try:
            fresh = await db.reports.find_one({"id": report_id}, {"_id": 0})
            await _send_customer_pdf(fresh.get("report_type"), fresh)
        except Exception as e:
            logger.error(f"customer email failed: {e}")
    return {"ok": True, "status": status}


@api_router.patch("/reports/{report_id}/signature")
async def add_report_signature(report_id: str, payload: dict, user: dict = Depends(get_current_user)):
    key = payload.get("key")
    signature = payload.get("signature") or {}
    allowed = {
        "issuer": ("technician", "supervisor", "head_maintenance", "admin"),
        "customer": tuple(ROLES),
        "checker": ("supervisor", "head_maintenance", "admin"),
        "approver": ("head_maintenance", "admin"),
        "fujitec_rep": ("technician", "supervisor", "head_maintenance", "admin"),
    }
    if key not in allowed:
        raise HTTPException(status_code=400, detail="Invalid signature key")
    if user["role"] not in allowed[key]:
        raise HTTPException(status_code=403, detail="Anda tidak berwenang menandatangani bagian ini")
    signature.setdefault("signed_at", now_iso())
    if not signature.get("name"):
        signature["name"] = user["name"]
    res = await db.reports.update_one({"id": report_id}, {"$set": {f"signatures.{key}": signature, "updated_at": now_iso()}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


@api_router.delete("/reports/{report_id}")
async def delete_report(report_id: str, user: dict = Depends(get_current_user)):
    doc = await db.reports.find_one({"id": report_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    if user["role"] not in ("admin", "supervisor", "head_maintenance") and doc.get("technician_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    await db.reports.delete_one({"id": report_id})
    return {"ok": True}


async def _fetch_photos(file_ids):
    photos = {}
    for fid in set([f for f in file_ids if f]):
        rec = await db.files.find_one({"id": fid, "is_deleted": False})
        if rec:
            try:
                data, _ = get_object(rec["storage_path"])
                photos[fid] = data
            except Exception:
                pass
    return photos


@api_router.get("/reports/{report_id}/pdf")
async def report_pdf(report_id: str, auth: str = Query(None)):
    require_pdf_token(auth, report_id)
    doc = await db.reports.find_one({"id": report_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    rtype = doc.get("report_type")
    if rtype == "ECR":
        photos = await _fetch_photos(doc.get("photo_file_ids", []) or [])
        pdf_bytes = build_ecr_pdf(doc, photos=photos)
        prefix, unit = "ECR", doc.get("unit_no")
    else:
        ids = []
        for p in doc.get("parts_replaced", []) or []:
            ids += [p.get("before_photo_file_id"), p.get("after_photo_file_id")]
        photos = await _fetch_photos(ids)
        pdf_bytes = build_hor_pdf(doc, photos=photos)
        prefix, unit = "HOR", doc.get("lift_number")
    filename = std_filename(prefix, doc.get("report_date"), doc.get("job_number"),
                            doc.get("site_name"), unit)
    return StreamingResponse(iter([pdf_bytes]), media_type="application/pdf",
                             headers={"Content-Disposition": f'inline; filename="{filename}"'})



# ---------------------------------------------------------------------------
# PDF secure tokens, customer email, unit history, excel, daily summary
# ---------------------------------------------------------------------------
@api_router.get("/inspections/{inspection_id}/pdf-token")
async def inspection_pdf_token(inspection_id: str, user: dict = Depends(get_current_user)):
    if not await db.inspections.find_one({"id": inspection_id}, {"_id": 1}):
        raise HTTPException(status_code=404, detail="Not found")
    return {"token": create_pdf_token(inspection_id)}


@api_router.get("/reports/{report_id}/pdf-token")
async def report_pdf_token(report_id: str, user: dict = Depends(get_current_user)):
    if not await db.reports.find_one({"id": report_id}, {"_id": 1}):
        raise HTTPException(status_code=404, detail="Not found")
    return {"token": create_pdf_token(report_id)}


async def _resolve_customer_email(doc):
    em = (doc.get("customer_email") or "").strip()
    if em:
        return em
    jn = doc.get("job_number")
    if jn:
        b = await db.buildings.find_one({"job_number": {"$regex": f"^{jn}$", "$options": "i"}})
        if b and b.get("customer_email"):
            return b["customer_email"]
    return ""


async def _send_customer_pdf(kind, doc):
    if not doc:
        return
    email = await _resolve_customer_email(doc)
    if not email:
        return
    if kind == "SIR":
        ids = [p.get("photo_file_id") for it in doc.get("items", []) for p in it.get("points", [])]
        pdf = build_sir_pdf(doc, photos=await _fetch_photos(ids))
        fn = std_filename("SIR", doc.get("inspection_date"), doc.get("job_number"), doc.get("site_name"), doc.get("lift_number"))
        label = "Service Inspection Report (SIR)"
    elif kind == "ECR":
        pdf = build_ecr_pdf(doc, photos=await _fetch_photos(doc.get("photo_file_ids", []) or []))
        fn = std_filename("ECR", doc.get("report_date"), doc.get("job_number"), doc.get("site_name"), doc.get("unit_no"))
        label = "Call Back Report (ECR)"
    else:
        ids = []
        for p in doc.get("parts_replaced", []) or []:
            ids += [p.get("before_photo_file_id"), p.get("after_photo_file_id")]
        pdf = build_hor_pdf(doc, photos=await _fetch_photos(ids))
        fn = std_filename("HOR", doc.get("report_date"), doc.get("job_number"), doc.get("site_name"), doc.get("lift_number"))
        label = "Hand Over Report (HOR)"
    attachments = [{"filename": fn, "content": base64.b64encode(pdf).decode()}]
    await send_email(email, f"[Fujitec] {label} - {doc.get('site_name','')} (Approved)",
                     approval_html(doc, label), attachments=attachments)


@api_router.get("/units/history")
async def unit_history(job_number: str = Query(...), user: dict = Depends(get_current_user)):
    rx = {"$regex": f"^{job_number}$", "$options": "i"}
    insp = await db.inspections.find({"job_number": rx}, {"_id": 0, "items": 0, "signatures": 0}).to_list(500)
    reps = await db.reports.find({"job_number": rx}, {"_id": 0}).to_list(500)
    events = []
    for d in insp:
        events.append({"kind": "SIR", "id": d["id"], "date": d.get("inspection_date"),
                       "status": d.get("status"), "technician": d.get("technician_name"),
                       "detail": f"{d.get('checklist_type','')} · {d.get('site_name','')} · {d.get('lift_number','')}",
                       "spare_parts": d.get("spare_parts", [])})
    for d in reps:
        events.append({"kind": d.get("report_type"), "id": d["id"], "date": d.get("report_date"),
                       "status": d.get("status"), "technician": d.get("technician_name"),
                       "detail": f"{d.get('site_name','')} · {d.get('unit_no') or d.get('lift_number','')}"})
    events.sort(key=lambda x: x.get("date") or "", reverse=True)
    building = await db.buildings.find_one({"job_number": rx}, {"_id": 0})
    return {"job_number": job_number, "building": building or {}, "events": events}


@api_router.get("/export/excel")
async def export_excel(user: dict = Depends(get_current_user)):
    import openpyxl
    import io as _io
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SIR"
    ws.append(["Job Number", "Site", "Lift", "Type", "Date", "Time", "Technician", "Status", "Spare Parts"])
    for d in await db.inspections.find({}, {"_id": 0}).to_list(5000):
        sp = "; ".join([f"{s.get('name','')} ({s.get('maintenance_status','')})" for s in d.get("spare_parts", [])])
        ws.append([d.get("job_number"), d.get("site_name"), d.get("lift_number"), d.get("checklist_type"),
                   d.get("inspection_date"), f"{d.get('time_from','')}-{d.get('time_to','')}",
                   d.get("technician_name"), d.get("status"), sp])
    ws2 = wb.create_sheet("ECR-HOR")
    ws2.append(["Type", "Job Number", "Site", "Unit", "Date", "Technician", "Status"])
    for d in await db.reports.find({}, {"_id": 0}).to_list(5000):
        ws2.append([d.get("report_type"), d.get("job_number"), d.get("site_name"),
                    d.get("unit_no") or d.get("lift_number"), d.get("report_date"),
                    d.get("technician_name"), d.get("status")])
    ws3 = wb.create_sheet("Spare Parts")
    ws3.append(["Spare Part", "Qty", "Site", "Job/Unit", "Sales", "Customer PO", "Inventory", "Maintenance"])
    for d in await db.inspections.find({"spare_parts.0": {"$exists": True}}, {"_id": 0}).to_list(5000):
        for s in d.get("spare_parts", []):
            ws3.append([s.get("name"), s.get("quantity"), d.get("site_name"),
                        f"{d.get('job_number','')}/{d.get('lift_number','')}", s.get("sales_status"),
                        s.get("customer_po_status"), s.get("inventory_status"), s.get("maintenance_status")])
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": 'attachment; filename="fujitec-reports.xlsx"'})


async def _pending_spare_rows():
    docs = await db.inspections.find({"spare_parts.0": {"$exists": True}}, {"_id": 0}).to_list(3000)
    rows = []
    for d in docs:
        for sp in d.get("spare_parts", []):
            if sp.get("maintenance_status") != "replaced":
                rows.append({**sp, "site_name": d.get("site_name"), "job_number": d.get("job_number"),
                             "lift_number": d.get("lift_number")})
    return rows


async def send_daily_summary():
    rows = await _pending_spare_rows()
    recipients = set()
    if OWNER_EMAIL:
        recipients.add(OWNER_EMAIL)
    async for u in db.users.find({"role": "head_maintenance"}, {"email": 1}):
        recipients.add(u["email"])
    html = daily_summary_html(rows)
    for r in recipients:
        await send_email(r, f"[Fujitec SIR] Ringkasan Harian — {len(rows)} spare part pending", html)
    return len(rows)


@api_router.post("/admin/daily-summary")
async def trigger_daily_summary(user: dict = Depends(require_roles("admin", "head_maintenance"))):
    n = await send_daily_summary()
    return {"ok": True, "pending": n}


# ---------------------------------------------------------------------------
# Dashboard stats
# ---------------------------------------------------------------------------
@api_router.get("/dashboard/stats")
async def dashboard_stats(user: dict = Depends(get_current_user)):
    query = _visible_query(user)
    docs = await db.inspections.find(query, {"_id": 0}).to_list(2000)
    total = len(docs)
    by_status = {"draft": 0, "submitted": 0, "checked": 0, "approved": 0}
    by_type = {"1M": 0, "3M": 0, "12M": 0}
    spare_total = 0
    spare_open = 0
    inv = {"pending": 0, "available": 0, "ordering": 0, "no_stock": 0, "discontinued": 0}
    recent = []
    for d in docs:
        by_status[d.get("status", "draft")] = by_status.get(d.get("status", "draft"), 0) + 1
        if d.get("checklist_type") in by_type:
            by_type[d["checklist_type"]] += 1
        for sp in d.get("spare_parts", []):
            spare_total += 1
            if sp.get("maintenance_status") != "replaced":
                spare_open += 1
            inv[sp.get("inventory_status", "pending")] = inv.get(sp.get("inventory_status", "pending"), 0) + 1
    for d in docs[:6]:
        recent.append({k: d.get(k) for k in
                       ("id", "job_number", "site_name", "lift_number", "checklist_type",
                        "status", "inspection_date", "technician_name")})
    recent.sort(key=lambda x: x.get("inspection_date", ""), reverse=True)
    return {"total_inspections": total, "by_status": by_status, "by_type": by_type,
            "spare_total": spare_total, "spare_open": spare_open,
            "inventory_breakdown": inv, "recent": recent[:6]}


# ---------------------------------------------------------------------------
# Startup
# ---------------------------------------------------------------------------
@app.on_event("startup")
async def startup():
    try:
        await db.users.create_index("email", unique=True)
        await db.login_attempts.create_index("identifier")
        await db.inspections.create_index("id", unique=True)
        await db.files.create_index("id", unique=True)
    except Exception as e:
        logger.warning(f"index creation: {e}")
    # Seed admin + demo users
    await seed_users()
    await load_buildings()
    try:
        # init_storage()
        pass
        # logger.info("Storage initialized")
    except Exception as e:
        # logger.error(f"Storage init failed: {e}")
        logger.warning(f"Storage init skipped: {e}")
    try:
        scheduler.add_job(send_daily_summary, "cron", hour=7, minute=0, id="daily_summary", replace_existing=True)
        if not scheduler.running:
            scheduler.start()
    except Exception as e:
        logger.error(f"scheduler start failed: {e}")


async def seed_users():
    seeds = [
        (os.environ.get("ADMIN_EMAIL", "admin@fujitec.com"), os.environ.get("ADMIN_PASSWORD", "admin123"), "Administrator", "admin"),
        ("teknisi@fujitec.com", "teknisi123", "Budi Teknisi", "technician"),
        ("supervisor@fujitec.com", "super123", "Sari Supervisor", "supervisor"),
        ("kepala@fujitec.com", "kepala123", "Head Maintenance", "head_maintenance"),
        ("sales@fujitec.com", "sales123", "Sales Service", "sales"),
        ("inventory@fujitec.com", "inv123", "Inventory Team", "inventory"),
        ("customer@fujitec.com", "cust123", "Customer PIC", "customer"),
    ]
    for email, pw, name, role in seeds:
        existing = await db.users.find_one({"email": email})
        if not existing:
            await db.users.insert_one({"email": email, "password_hash": hash_password(pw),
                                       "name": name, "role": role, "created_at": now_iso()})
        elif existing.get("password_hash") and not verify_password(pw, existing["password_hash"]):
            await db.users.update_one({"email": email}, {"$set": {"password_hash": hash_password(pw)}})
    # Ensure owner (Google) account exists as admin
    if OWNER_EMAIL and not await db.users.find_one({"email": OWNER_EMAIL}):
        await db.users.insert_one({"email": OWNER_EMAIL, "name": "Agus Supriyadi",
                                   "role": "admin", "auth_provider": "google", "created_at": now_iso()})


async def load_buildings():
    try:
        existing = await db.buildings.find_one({})
        if existing and "maintenance_period" in existing:
            return
        path = ROOT_DIR / "data_buildings.json"
        if not path.exists():
            return
        rows = json.loads(path.read_text(encoding="utf-8"))
        if rows:
            await db.buildings.delete_many({})
            await db.buildings.insert_many(rows)
            await db.buildings.create_index("job_number")
            logger.info(f"Loaded {len(rows)} buildings")
    except Exception as e:
        logger.error(f"load_buildings failed: {e}")


@app.on_event("shutdown")
async def shutdown():
    try:
        if scheduler.running:
            scheduler.shutdown(wait=False)
    except Exception:
        pass
    client.close()


app.include_router(api_router)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)
